"""
google_news_scraper.py
======================
Fetches Google News RSS for airline loyalty programs, credit card products,
and company-level news (M&A, earnings, bankruptcy) for the previous month.

Three query buckets per competitor:
  1. Loyalty / program changes
  2. Credit card product news
  3. Company-level news (M&A, earnings, bankruptcy)

Auto-detects the PREVIOUS calendar month.
Outputs: data/google_news_YYYY_MM.csv + data/google_news_YYYY_MM.xlsx (with Themes tab)

Run:         python google_news_scraper.py
Specific:    python google_news_scraper.py --year 2026 --month 3
No summary:  python google_news_scraper.py --no-summary
Test mode:   python google_news_scraper.py --test --no-summary
"""

import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..'))

import argparse
import calendar
import csv
import os
import re
import time
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import datetime, timezone
from urllib.parse import quote

import requests
from summarizer import summarize, check_api_key
from config import GOOGLE_NEWS_QUERIES, NEWSROOM_SOURCES


QUERY_BUCKETS = [
    ("Loyalty / Program",   "loyalty_query"),
    ("Credit Card Product", "product_query"),
    ("Company News",        "company_query"),
]

GOOGLE_NEWS_RSS = "https://news.google.com/rss/search?q={query}&hl=en-US&gl=US&ceid=US:en"
WEB_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")

STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "in", "for", "on", "at",
    "is", "its", "with", "as", "by", "from", "that", "this", "are", "was",
    "will", "be", "it", "but", "not", "has", "have", "had", "what", "how",
    "why", "who", "after", "before", "than", "more", "new", "into", "up",
    "could", "would", "should", "still", "here", "about", "over", "just",
    "says", "say", "said", "amid", "despite", "also", "now", "off", "air",
    "lines", "airline", "airlines", "one", "two", "three", "first", "last",
}

SOURCE_PRIORITY_GLOBAL = [
    "Reuters", "WSJ", "Bloomberg", "Bloomberg.com", "Bloomberg Law News",
    "Financial Times", "Financial News London", "AP News",
    "CNBC", "American Banker", "Business Wire", "PR Newswire",
    "PR Newswire Canada", "Payments Dive", "Banking Dive",
    "Skift", "Simple Flying", "FlightGlobal", "Airline Weekly",
    "AirlineGeeks.com", "AirInsight", "The Air Current",
    "Travel Weekly", "Aviation Week", "ch-aviation",
    "Aircraft Interiors International", "Centre for Aviation",
    "Business Travel Executive", "AeroTime",
    "The Points Guy", "NerdWallet", "Forbes", "Upgraded Points",
    "One Mile at a Time", "Frequent Miler", "viewfromthewing.com",
    "Thrifty Traveler", "Live From A Lounge", "Bankrate",
    "The Motley Fool", "Kiplinger", "Investopedia",
    "Eye of the Flyer", "CardRates.com",
    "MarketWatch", "Barron's", "Fortune", "Fast Company",
    "Axios", "The Washington Post", "The New York Times",
    "USA Today", "Business Insider", "The Boston Globe",
    "Seeking Alpha", "Yahoo Finance", "Investing.com",
    "PYMNTS.com", "Morning Consult",
    "Alaska Airlines", "American Airlines", "Delta News Hub",
    "JPMorganChase", "Chase Bank", "American Express",
    "Citigroup", "JetBlue Airways Corporation - Investor Relations",
    "Frontier Airlines",
]

GUIDE_KEYWORDS = [
    "how to", "best way", "best ways", "guide to", "guide:",
    "complete guide", "step by step", "step-by-step",
    "best credit cards", "best cards", "top cards",
    "ways to earn", "ways to redeem", "ways to fly",
    "ways to get", "ways to book", "ways to use",
    "what is", "what are", "everything you need",
    "should you", "is it worth", "worth it?",
    "vs.", " vs ", "comparison", "comparing",
    "review:", "review —", "card review",
    "explained", "here's why", "here's how",
    "tips for", "tips to", "tricks to",
    "beginners guide", "ultimate guide",
    "checklist", "cheat sheet", "breakdown",
    "ranked", "ranking", "best of",
    "best time to apply", "smart ways", "my strategy",
    "i never ", "here's what i do", "here's what to do",
    "loyalty program review", "class review", "business class review",
    "polaris", "review [",
    "ways to save", "things to know", "perks worth",
    "benefits of", "how i saved", "cost me $",
    "wrong credit card", "card benefits",
    "things you should", "reasons to",
    "who qualifies", "do you qualify", "see if you qualify",
    "what to know", "what customers should know",
    "eligibility", "here's who qualifies",
    "when to expect", "what to expect",
    "stocks to watch", "final trades",
    "wall street expects", "earnings growth",
    "gear up for", "on deck",
    "buy the dip", "price target",
    "analyst questions", "earnings call",
    "valuation after", "share performance",
    "reiterates buy", "raises price",
    "workaround", "this trick", "use this instead",
    "always keep", "i always", "my favorite",
    "after you die", "simple mistake", "cost me",
    "end of year", "ways i look", "save thousands",
    "how i book", "summer travel", "is a premium",
    "perks to justify", "beginner rewards",
    "i've covered", "i covered", "could save you",
    "these 3 cards", "these 2 cards", "traveling abroad",
    "refer a friend", "best annual fee", "best capital one",
    "prices are spiking", "pays warren buffett", "stock pays",
    "moved up by", "stock storms", "fallen 23",
    "wall street forecasters", "stock is climbing",
    "bullish or bearish", "stock rallies", "stock outlook",
    "off lows", "ahead of q", "stock soars", "stock jumps",
    "too late to", "stock takes off", "savings accounts",
]


def get_target_month(year=None, month=None):
    if year and month:
        return int(year), int(month)
    today = datetime.now(timezone.utc)
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def parse_date(s):
    if not s:
        return None
    for fmt in ["%a, %d %b %Y %H:%M:%S %z", "%a, %d %b %Y %H:%M:%S GMT",
                "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d"]:
        try:
            dt = datetime.strptime(s.strip()[:35], fmt)
            return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt
        except ValueError:
            continue
    return None


def in_target_month(dt, year, month):
    if dt is None:
        return False
    dt = dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt
    return dt.year == year and dt.month == month


def fetch_google_news(company, category_label, query, year, month):
    last_day   = calendar.monthrange(year, month)[1]
    full_query = f"{query} after:{year}-{month:02d}-01 before:{year}-{month:02d}-{last_day}"
    url = GOOGLE_NEWS_RSS.format(query=quote(full_query))

    results = []
    try:
        resp = requests.get(url, headers=WEB_HEADERS, timeout=15)
        resp.raise_for_status()
        root  = ET.fromstring(resp.content)
        items = root.findall(".//item")

        for item in items:
            title   = (item.findtext("title")   or "").strip()
            link    = (item.findtext("link")    or "").strip()
            pub_str = (item.findtext("pubDate") or "").strip()
            src_el  = item.find("source")
            source  = src_el.text.strip() if src_el is not None else "Google News"
            pub_dt  = parse_date(pub_str)

            if not in_target_month(pub_dt, year, month):
                continue

            results.append({
                "company":  company,
                "category": category_label,
                "source":   source,
                "date":     pub_dt.strftime("%Y-%m-%d") if pub_dt else "",
                "title":    title,
                "url":      link,
            })

        time.sleep(0.8)
    except Exception as e:
        print(f"      ❌ Error: {e}")

    return results




def fetch_newsroom_articles(year, month, sources):
    """Fetch newsroom/press release articles and normalize to google_news format."""
    import calendar as _cal
    results = []
    for newsroom in sources:
        last_day   = _cal.monthrange(year, month)[1]
        full_query = f"{newsroom['query']} after:{year}-{month:02d}-01 before:{year}-{month:02d}-{last_day}"
        url = GOOGLE_NEWS_RSS.format(query=quote(full_query))
        try:
            resp = requests.get(url, headers=WEB_HEADERS, timeout=15)
            resp.raise_for_status()
            root  = ET.fromstring(resp.content)
            items = root.findall(".//item")
            count = 0
            for item in items:
                title   = (item.findtext("title")   or "").strip()
                link    = (item.findtext("link")    or "").strip()
                pub_str = (item.findtext("pubDate") or "").strip()
                src_el  = item.find("source")
                source  = src_el.text.strip() if src_el is not None else "Google News"
                pub_dt  = parse_date(pub_str)
                if not in_target_month(pub_dt, year, month):
                    continue
                # Normalize to google_news field names
                results.append({
                    "company":  newsroom["company"],
                    "category": "Company News",  # newsroom = company announcements
                    "source":   source,
                    "date":     pub_dt.strftime("%Y-%m-%d") if pub_dt else "",
                    "title":    title,
                    "url":      link,
                })
                count += 1
            time.sleep(0.8)
            print(f"    📰 Newsroom {newsroom['company']}: {count} articles")
        except Exception as e:
            print(f"    ❌ Newsroom error {newsroom['company']}: {e}")
    return results

def generate_theme_name(cluster_rows, max_words=8):
    """Generate a theme name from cluster headlines using word frequency."""
    word_counts = Counter()
    for row in cluster_rows:
        clean = re.sub(r"[^a-z0-9 ]", "", row["title"].lower())
        clean = re.sub(r" - [a-z0-9 ]+$", "", clean)
        words = [w for w in clean.split() if w not in STOPWORDS and len(w) > 2]
        word_counts.update(words)
    top_words = [w for w, _ in word_counts.most_common(max_words)]
    return " ".join(top_words).title()


def filter_guides(rows):
    """Pass 0 — remove evergreen guides, reviews, and listicles. Keep news only."""
    filtered = []
    for row in rows:
        title_lower = row["title"].lower()
        if any(kw in title_lower for kw in GUIDE_KEYWORDS):
            continue
        filtered.append(row)
    removed = len(rows) - len(filtered)
    print(f"\n  🗞️  Guide filter: {len(rows)} → {len(filtered)} news articles ({removed} guides removed)")
    return filtered


def normalize_title(title):
    t = re.sub(r"[^a-z0-9 ]", "", title.lower())
    return " ".join(t.split()[:8])


def deduplicate_exact(rows):
    """Pass 1 — remove exact title matches."""
    seen, deduped = {}, []
    for row in rows:
        key = (row["company"], normalize_title(row["title"]))
        if key not in seen:
            seen[key] = True
            deduped.append(row)
    print(f"\n  🔄 Exact dedup: {len(rows)} → {len(deduped)} unique")
    return deduped


def deduplicate_semantic(rows, threshold=0.88):
    """Pass 2 — remove semantically similar articles (same story, different headline)."""
    try:
        from sentence_transformers import SentenceTransformer
        import numpy as np
    except ImportError:
        print("  ⚠️  sentence-transformers not installed — skipping semantic dedup")
        return rows

    if len(rows) < 2:
        return rows

    print(f"  🧠 Running semantic dedup on {len(rows)} articles (threshold: {threshold})...")
    model = SentenceTransformer('all-MiniLM-L6-v2')
    texts = [r['title'] for r in rows]
    embeddings = model.encode(texts, show_progress_bar=False)

    import numpy as np
    norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
    embeddings = embeddings / np.where(norms == 0, 1, norms)
    similarity = np.dot(embeddings, embeddings.T)

    def source_rank(row):
        for i, s in enumerate(SOURCE_PRIORITY_GLOBAL):
            if s.lower() in row.get("source", "").lower():
                return i
        return len(SOURCE_PRIORITY_GLOBAL)

    kept = []
    dropped = set()

    for i, row in enumerate(rows):
        if i in dropped:
            continue
        kept.append(row)
        for j in range(i + 1, len(rows)):
            if j not in dropped and similarity[i][j] >= threshold:
                if source_rank(rows[j]) < source_rank(row):
                    kept[-1] = rows[j]
                dropped.add(j)

    print(f"  🧠 Semantic dedup: {len(rows)} → {len(kept)} unique")
    return kept


def source_diversity_filter(rows, window_days=7, eps=0.20, min_samples=2, eps_by_category=None):
    """Pass 3 — DBSCAN clustering within company + date windows.
    For each cluster keeps the best-ranked source, saves cluster as a theme.
    Noise points (label=-1) are kept as-is since they are unique articles.
    eps_by_category: dict mapping category name to eps override e.g. {"Credit Card Product": 0.25}
    Returns (kept_articles, themes)."""
    if eps_by_category is None:
        eps_by_category = {}
    try:
        from sentence_transformers import SentenceTransformer
        from sklearn.cluster import DBSCAN
        import numpy as np
    except ImportError:
        print("  ⚠️  sentence-transformers or scikit-learn not installed — skipping DBSCAN")
        # Fallback: keep all, no themes
        return rows, []

    def source_rank(row):
        for i, s in enumerate(SOURCE_PRIORITY_GLOBAL):
            if s.lower() in row.get("source", "").lower():
                return i
        return len(SOURCE_PRIORITY_GLOBAL)

    def parse_row_date(row):
        try:
            return datetime.strptime(row["date"], "%Y-%m-%d")
        except:
            return None

    def make_theme(cluster_rows):
        best = sorted(cluster_rows, key=source_rank)[0]
        all_dates = sorted([r["date"] for r in cluster_rows if r["date"]])
        date_range = (all_dates[0] if all_dates[0] == all_dates[-1]
                      else f"{all_dates[0]} to {all_dates[-1]}")
        return {
            "company":       best["company"],
            "category":      best["category"],
            "theme":         generate_theme_name(cluster_rows),
            "article_count": len(cluster_rows),
            "date_range":    date_range,
            "best_headline": best["title"],
            "best_source":   best["source"],
            "best_url":      best["url"],
            "all_sources":   " | ".join(r["source"] for r in cluster_rows),
            "all_urls":      " | ".join(r["url"] for r in cluster_rows),
        }

    print(f"  🔬 DBSCAN clustering (eps={eps}, min_samples={min_samples}, window={window_days}d)...")
    model = SentenceTransformer('all-MiniLM-L6-v2')

    kept   = []
    themes = []
    companies = sorted(set(r["company"] for r in rows))

    for company in companies:
        company_rows = [r for r in rows if r["company"] == company]
        if not company_rows:
            continue

        company_rows_sorted = sorted(company_rows, key=lambda r: (r["date"], source_rank(r)))

        # Build date windows greedily — anchor on each unassigned article
        windowed_groups = []
        row_used = set()

        for i, anchor in enumerate(company_rows_sorted):
            if i in row_used:
                continue
            anchor_date = parse_row_date(anchor)
            window = [i]
            for j in range(i + 1, len(company_rows_sorted)):
                if j in row_used:
                    continue
                other_date = parse_row_date(company_rows_sorted[j])
                if anchor_date and other_date:
                    if abs((other_date - anchor_date).days) <= window_days:
                        window.append(j)
            windowed_groups.append(window)
            for idx in window:
                row_used.add(idx)

        # Run DBSCAN within each window
        for window_indices in windowed_groups:
            window_rows = [company_rows_sorted[idx] for idx in window_indices]

            if len(window_rows) == 1:
                kept.append(window_rows[0])
                continue

            texts = [r["title"] for r in window_rows]
            embeddings = model.encode(texts, show_progress_bar=False)
            norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
            embeddings = embeddings / np.where(norms == 0, 1, norms)
            distance_matrix = np.clip(1 - np.dot(embeddings, embeddings.T), 0, None)

            # Use category-specific eps if defined, else default
            category = window_rows[0].get("category", "")
            effective_eps = eps_by_category.get(category, eps)
            db = DBSCAN(eps=effective_eps, min_samples=min_samples, metric="precomputed")
            labels = db.fit_predict(distance_matrix)

            cluster_map = {}
            for idx, label in enumerate(labels):
                cluster_map.setdefault(label, []).append(idx)

            for label, indices in cluster_map.items():
                cluster_rows = [window_rows[idx] for idx in indices]

                if label == -1:
                    # Noise — unique articles, keep all
                    kept.extend(cluster_rows)
                else:
                    # Real cluster — keep best source only
                    best = sorted(cluster_rows, key=source_rank)[0]
                    kept.append(best)
                    if len(cluster_rows) >= 2:
                        themes.append(make_theme(cluster_rows))

    print(f"  📰 DBSCAN clustering: {len(rows)} → {len(kept)} articles")
    print(f"  🗂️  Themes generated: {len(themes)}")
    return kept, themes



THEME_SUMMARY_PROMPT = (
    "You are a competitive intelligence analyst for a travel credit card product team.\n\n"
    "Competitor: {company}\n"
    "Category: {category}\n"
    "Theme: {theme}\n"
    "Article text:\n{headlines}\n\n"
    "Write ONE sentence (max 30 words) summarizing the competitive intelligence implication.\n"
    "Format: \"[Competitor] [did X], [competitive implication for card products].\"\n"
    "Return only the sentence. No preamble, no quotes."
)


def summarize_themes(themes, api_key):
    """Summarize each theme cluster using Claude. Much cheaper than per-article summaries."""
    if not api_key:
        print("  ⚠️  ANTHROPIC_API_KEY not set — skipping theme summaries")
        return themes

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
    except Exception as e:
        print(f"  ⚠️  Claude init failed: {e}")
        return themes

    def _fetch_article_text(url, max_chars=4000):
        """Fetch real article text — follows Google News RSS redirects to actual article."""
        try:
            import requests as _req
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
            }
            # Follow redirects — Google News RSS URLs redirect to the real article
            resp = _req.get(url, headers=headers, timeout=12, allow_redirects=True)
            resp.raise_for_status()

            # Check if we got actual content or JavaScript redirect
            content_type = resp.headers.get("content-type", "")
            if "javascript" in content_type or resp.text.strip().startswith("var "):
                return ""

            text = re.sub(r"<[^>]+>", " ", resp.text)
            text = re.sub(r"&[a-z]+;", " ", text)
            text = re.sub(r"\s+", " ", text).strip()
            # Skip if result looks like JS code
            if text.strip().startswith(("var ", "function ", "window.", "(function")):
                return ""
            return text[:max_chars]
        except:
            return ""

    print(f"\n  🤖 Summarizing {len(themes)} themes with Claude...")
    for i, theme in enumerate(themes, 1):
        # Try to fetch real article text — fall back to headline if fetch fails
        article_text = _fetch_article_text(theme["best_url"])
        if not article_text:
            article_text = f"Theme: {theme['theme']}. Headline: {theme['best_headline']}. Date: {theme['date_range']}."

        prompt = THEME_SUMMARY_PROMPT.format(
            company=theme["company"],
            category=theme["category"],
            theme=theme["theme"],
            headlines=article_text,
        )
        try:
            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=80,
                messages=[{"role": "user", "content": prompt}]
            )
            theme["summary"] = response.content[0].text.strip().strip('"').strip("'")
            print(f"    [{i}/{len(themes)}] {theme['company']} | {theme['theme'][:50]}...")
        except Exception as e:
            print(f"    ⚠️  Failed: {e}")
            theme["summary"] = ""
        time.sleep(0.3)

    return themes

FIELDNAMES = ["company", "category", "date", "title", "summary", "source", "url"]

THEME_FIELDNAMES = [
    "company", "category", "theme", "article_count", "date_range",
    "best_headline", "best_source", "best_url", "all_sources", "all_urls",
]


def save_csv(rows, year, month, test_mode=False):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    suffix = "_test" if test_mode else ""
    filename = os.path.join(OUTPUT_DIR, f"news_articles_{year}_{month:02d}{suffix}.csv")
    rows_sorted = sorted(rows, key=lambda r: (r["company"], r["category"], r["date"]))
    with open(filename, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader(); w.writerows(rows_sorted)
    print(f"\n  💾 Saved {len(rows_sorted)} rows → {filename}")
    return filename


def save_xlsx(rows, themes, year, month, test_mode=False):
    """Save articles + themes as a two-sheet Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️  openpyxl not installed — skipping Excel output")
        return None

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    suffix = "_test" if test_mode else ""
    filename = os.path.join(OUTPUT_DIR, f"news_themes_{year}_{month:02d}{suffix}.xlsx")

    wb = openpyxl.Workbook()

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="1F3864")
    theme_fill   = PatternFill("solid", start_color="1F4E79")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", start_color="F5F8FF")

    # ── Sheet 1: Articles ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Articles"

    art_headers = [
        ("Company", 16), ("Category", 20), ("Date", 12),
        ("Title", 60), ("Summary", 50), ("Source", 22), ("URL", 40),
    ]
    for col, (h, w) in enumerate(art_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 28
    ws1.freeze_panes = "A2"

    rows_sorted = sorted(rows, key=lambda r: (r["company"], r["category"], r["date"]))
    for row_idx, row in enumerate(rows_sorted, 2):
        is_alt = row_idx % 2 == 0
        values = [
            row.get("company", ""), row.get("category", ""), row.get("date", ""),
            row.get("title", ""), row.get("summary", ""),
            row.get("source", ""), row.get("url", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = center if col in (1, 2, 3, 6) else left_wrap
            if is_alt:
                cell.fill = alt_fill
        ws1.row_dimensions[row_idx].height = 32

    ws1.auto_filter.ref = f"A1:G1"

    # ── Sheet 2: Themes ────────────────────────────────────────────
    ws2 = wb.create_sheet("Themes")

    theme_headers = [
        ("Company", 16), ("Category", 20), ("Theme", 45),
        ("# Articles", 12), ("Date Range", 22),
        ("Summary", 60),
        ("Best Headline", 55), ("Best Source", 22), ("Best URL", 40),
        ("All Sources", 50), ("All URLs", 80),
    ]
    for col, (h, w) in enumerate(theme_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = theme_fill
        cell.alignment = center; cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    themes_sorted = sorted(themes, key=lambda t: (t["company"], t["category"], -t["article_count"]))
    high_fill = PatternFill("solid", start_color="FFF2CC")  # yellow for 5+ articles
    med_fill  = PatternFill("solid", start_color="DDEEFF")  # blue for 3-4 articles

    for row_idx, theme in enumerate(themes_sorted, 2):
        count = theme["article_count"]
        row_fill = high_fill if count >= 5 else (med_fill if count >= 3 else None)
        values = [
            theme["company"], theme["category"], theme["theme"],
            count, theme["date_range"],
            theme.get("summary", ""),
            theme["best_headline"], theme["best_source"], theme["best_url"],
            theme["all_sources"], theme["all_urls"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = center if col in (1, 2, 4, 5, 8) else left_wrap
            if row_fill:
                cell.fill = row_fill
        ws2.row_dimensions[row_idx].height = 36

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(theme_headers))}1"

    wb.save(filename)
    print(f"  📊 Excel saved → {filename} (Articles: {len(rows_sorted)}, Themes: {len(themes_sorted)})")
    return filename


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--year",       type=int)
    parser.add_argument("--month",      type=int)
    parser.add_argument("--no-summary", action="store_true")
    parser.add_argument("--test",       action="store_true",
                        help="Test mode: 2 competitors only, no API costs")
    parser.add_argument("--threshold",  type=float, default=0.88,
                        help="Semantic similarity threshold (default: 0.88)")
    args = parser.parse_args()

    if args.test:
        from config_test import GOOGLE_NEWS_QUERIES_TEST
        queries = GOOGLE_NEWS_QUERIES_TEST
        print("\n  ⚠️  TEST MODE — using 2 competitors, output saved with _test suffix")
    else:
        queries = GOOGLE_NEWS_QUERIES

    year, month  = get_target_month(args.year, args.month)
    api_ok, msg  = check_api_key()
    do_summarize = not args.no_summary and api_ok

    if not args.no_summary and not api_ok:
        print(f"  ℹ️  {msg}")

    print(f"\n{'='*60}")
    print(f"  Google News Scraper  |  Target: {year}-{month:02d}")
    print(f"  Competitors: {len(queries)}  |  Buckets: {len(QUERY_BUCKETS)} each")
    print(f"  Summaries: {'✅ Claude' if do_summarize else '⏭️  skipped'}")
    print(f"  Semantic dedup threshold: {args.threshold}")
    print(f"{'='*60}")

    all_articles = []
    for competitor in queries:
        company = competitor["company"]
        print(f"\n  🔍 {company}")
        for bucket_label, bucket_key in QUERY_BUCKETS:
            results = fetch_google_news(company, bucket_label, competitor[bucket_key], year, month)
            print(f"    [{bucket_label}]: {len(results)}")
            all_articles.extend(results)

    # Fetch and merge newsroom/press release articles
    newsroom_sources = NEWSROOM_SOURCES if not args.test else NEWSROOM_SOURCES[:2]
    print(f"\n  📰 Fetching newsroom articles ({len(newsroom_sources)} sources)...")
    newsroom_articles = fetch_newsroom_articles(year, month, newsroom_sources)
    print(f"  📰 Newsroom total: {len(newsroom_articles)} articles")
    all_articles.extend(newsroom_articles)
    print(f"\n  📊 Combined total before dedup: {len(all_articles)}")

    # Four-pass deduplication
    all_articles          = filter_guides(all_articles)
    deduped               = deduplicate_exact(all_articles)
    deduped               = deduplicate_semantic(deduped, threshold=args.threshold)
    deduped, themes       = source_diversity_filter(deduped, window_days=7, eps=0.20, min_samples=2, eps_by_category={"Credit Card Product": 0.25, "Loyalty / Program": 0.25})

    # Summarize themes only (much cheaper than per-article)
    if do_summarize:
        themes = summarize_themes(themes, os.environ.get("ANTHROPIC_API_KEY", ""))
    else:
        for t in themes:
            t["summary"] = ""

    # Articles never get individual summaries — use theme summaries instead
    for a in deduped:
        a["summary"] = ""

    save_csv(deduped, year, month, test_mode=args.test)
    save_xlsx(deduped, themes, year, month, test_mode=args.test)

    by_cat = {}
    for a in deduped:
        by_cat[a["category"]] = by_cat.get(a["category"], 0) + 1

    print(f"\n{'='*60}")
    print(f"  ✅ Done!")
    for cat, count in sorted(by_cat.items()):
        print(f"     {cat}: {count}")
    print(f"     Themes: {len(themes)}")
    print(f"     Summaries: {sum(1 for a in deduped if a['summary'])}/{len(deduped)}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
