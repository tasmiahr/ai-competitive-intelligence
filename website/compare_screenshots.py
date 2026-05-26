"""
compare_screenshots.py
Compares current month screenshots against previous month using:
1. Pillow pixel diff — free gatekeeper, skips Claude if no change
2. Crops to bounding box of changed region — minimizes Claude Vision tokens
3. Claude Vision — only called when real changes detected
Writes change report to data/change_report.xlsx and data/change_summary.md
"""

import os
import sys
import json
import base64
import glob
import re
import requests
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageChops
import io

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
MODEL = "claude-sonnet-4-20250514"  # Sonnet not Opus — cheaper, still accurate
CHANGE_REPORT_XLSX = "data/change_report.xlsx"
CHANGE_SUMMARY_MD  = "data/change_summary.md"

# Pixel diff thresholds
PIXEL_DIFF_THRESHOLD = 500    # min changed pixels to trigger Claude
CROP_PADDING         = 50     # padding around bounding box when cropping


COMPARISON_PROMPT = """You are a senior competitive intelligence analyst comparing two cropped screenshots of an airline credit card webpage. The images show only the region where a change was detected.

FIRST IMAGE = previous month. SECOND IMAGE = current month.

Focus on what changed in this region. Check for:
- Welcome bonus miles/points amounts
- Annual fee changes
- Spend threshold or time period changes
- New or removed limited-time offers
- CTA button text or color changes
- Card reordering or new cards added
- Copy or headline changes
- Visual/design changes

For EACH change found, classify business impact:
- high: affects conversion or competitive position (bonus change, fee change, new offer)
- medium: affects messaging (copy update, CTA change, card reorder)
- low: cosmetic (footnote, small visual adjustment)

Return ONLY a JSON object:
{
  "has_changes": true or false,
  "change_summary": "2-3 sentence summary or 'No significant changes detected'",
  "changes": [
    {
      "card_name": "Card name or 'Page'",
      "section": "hero|card_offer|navigation|footer|cta|layout|copy|visual|promotional|other",
      "change_type": "bonus_increase|bonus_decrease|fee_change|earn_rate_change|new_offer|offer_removed|copy_change|headline_change|cta_change|layout_change|visual_change|card_reorder|new_badge|urgency_added|other",
      "previous_value": "What was there before",
      "current_value": "What is there now",
      "severity": "high|medium|low",
      "notes": "Competitive implication"
    }
  ],
  "confidence": "high|medium|low",
  "confidence_notes": "Any limitations"
}
Return ONLY the JSON. No markdown, no preamble."""


# ─────────────────────────────────────────────
# PIXEL DIFF GATEKEEPER
# ─────────────────────────────────────────────

def pixel_diff_check(prev_path: str, curr_path: str):
    """
    Compare two screenshots using Pillow pixel diff.
    Returns (has_change, cropped_prev_bytes, cropped_curr_bytes, pixel_count)
    If no change: returns (False, None, None, 0)
    If changed: returns (True, cropped_prev, cropped_curr, pixel_count)
    """
    try:
        img1 = Image.open(prev_path).convert("RGB")
        img2 = Image.open(curr_path).convert("RGB")

        # Resize to same dimensions if different
        if img1.size != img2.size:
            img2 = img2.resize(img1.size, Image.LANCZOS)

        diff = ImageChops.difference(img1, img2)
        bbox = diff.getbbox()

        if bbox is None:
            return False, None, None, 0

        # Count changed pixels
        diff_array = list(diff.getdata())
        changed_pixels = sum(1 for r, g, b in diff_array if r + g + b > 30)

        if changed_pixels < PIXEL_DIFF_THRESHOLD:
            return False, None, None, changed_pixels

        # Add padding to bounding box
        w, h = img1.size
        x1 = max(0, bbox[0] - CROP_PADDING)
        y1 = max(0, bbox[1] - CROP_PADDING)
        x2 = min(w, bbox[2] + CROP_PADDING)
        y2 = min(h, bbox[3] + CROP_PADDING)
        crop_box = (x1, y1, x2, y2)

        # Crop both images to changed region
        cropped_prev = img1.crop(crop_box)
        cropped_curr = img2.crop(crop_box)

        # Convert to bytes for API
        def to_bytes(img):
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return buf.getvalue()

        print(f"  🔍 Pixel diff: {changed_pixels} changed pixels, bbox={crop_box}")
        return True, to_bytes(cropped_prev), to_bytes(cropped_curr), changed_pixels

    except Exception as e:
        print(f"  ⚠️  Pixel diff error: {e} — falling back to full image comparison")
        return True, None, None, -1


def encode_image(image_path: str) -> tuple[bytes, str]:
    """Read image file, return (bytes, media_type)."""
    with open(image_path, "rb") as f:
        data = f.read()
    media_type = "image/png" if image_path.lower().endswith(".png") else "image/jpeg"
    return data, media_type


def compare_images(prev_path: str, curr_path: str) -> dict:
    """
    Full comparison pipeline:
    1. Pixel diff check — skip if no change
    2. Crop to changed region
    3. Send cropped images to Claude
    """
    # Step 1: Pixel diff gatekeeper
    has_change, prev_bytes, curr_bytes, pixel_count = pixel_diff_check(prev_path, curr_path)

    if not has_change:
        print(f"  ✅ No pixel change detected ({pixel_count} pixels) — skipping Claude")
        return {
            "has_changes": False,
            "change_summary": "No pixel changes detected — pages appear identical",
            "changes": [],
            "confidence": "high",
            "confidence_notes": f"Pixel diff: {pixel_count} changed pixels below threshold",
            "skipped_claude": True,
        }

    # Step 2: Use cropped images if available, else full images
    if prev_bytes and curr_bytes:
        prev_data = base64.standard_b64encode(prev_bytes).decode("utf-8")
        curr_data = base64.standard_b64encode(curr_bytes).decode("utf-8")
        prev_media = "image/png"
        curr_media = "image/png"
        print(f"  📐 Sending cropped region to Claude ({pixel_count} changed pixels)")
    else:
        prev_raw, prev_media = encode_image(prev_path)
        curr_raw, curr_media = encode_image(curr_path)
        prev_data = base64.standard_b64encode(prev_raw).decode("utf-8")
        curr_data = base64.standard_b64encode(curr_raw).decode("utf-8")
        print(f"  📐 Sending full images to Claude (pixel diff fallback)")

    # Step 3: Send to Claude
    response = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": MODEL,
            "max_tokens": 2000,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "PREVIOUS MONTH screenshot:"},
                        {"type": "image", "source": {"type": "base64", "media_type": prev_media, "data": prev_data}},
                        {"type": "text", "text": "CURRENT MONTH screenshot:"},
                        {"type": "image", "source": {"type": "base64", "media_type": curr_media, "data": curr_data}},
                        {"type": "text", "text": COMPARISON_PROMPT},
                    ],
                }
            ],
        },
        timeout=90,
    )

    if response.status_code != 200:
        return {
            "has_changes": False,
            "change_summary": f"API error {response.status_code}",
            "changes": [],
            "confidence": "low",
            "confidence_notes": response.text[:200],
        }

    raw = response.json()["content"][0]["text"].strip()
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    try:
        result = json.loads(raw)
        result["pixel_count"] = pixel_count
        return result
    except json.JSONDecodeError:
        return {
            "has_changes": False,
            "change_summary": "Parse error",
            "changes": [],
            "confidence": "low",
            "confidence_notes": raw[:300],
        }


def find_screenshot_folders() -> tuple[str | None, str | None]:
    folders = sorted(glob.glob("screenshots_*"), reverse=True)
    if len(folders) < 2:
        return None, folders[0] if folders else None
    return folders[1], folders[0]


def match_screenshots(prev_folder: str, curr_folder: str):
    def get_site_map(folder):
        files = glob.glob(f"{folder}/*_full.png") + glob.glob(f"{folder}/*_full.jpg")
        return {Path(f).stem.replace("_full", ""): f for f in files}

    prev_map = get_site_map(prev_folder)
    curr_map = get_site_map(curr_folder)
    pairs = []
    for site_key in sorted(set(prev_map) | set(curr_map)):
        site_name = site_key.replace("_", " ")
        pairs.append((site_name, prev_map.get(site_key), curr_map.get(site_key)))
    return pairs


SEVERITY_COLOR = {"high": "FF4444", "medium": "FF9900", "low": "00AA44"}

CHANGE_TYPE_LABELS = {
    "bonus_increase": "🟢 Bonus Increase", "bonus_decrease": "🔴 Bonus Decrease",
    "fee_change": "🟠 Fee Change", "earn_rate_change": "🔵 Earn Rate Change",
    "new_offer": "🟡 New Offer", "offer_removed": "⬜ Offer Removed",
    "copy_change": "✏️ Copy Change", "headline_change": "📢 Headline Change",
    "cta_change": "🔘 CTA Change", "layout_change": "📐 Layout Change",
    "card_reorder": "🔀 Card Reorder", "new_element": "➕ New Element",
    "visual_change": "🎨 Visual Change", "new_badge": "🏷️ New Badge",
    "urgency_added": "⚡ Urgency Added", "other": "⚪ Other",
}


def write_change_report(results, run_date, prev_folder, curr_folder):
    os.makedirs("data", exist_ok=True)
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F3864")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_sum["A1"] = "Competitive Intelligence: Visual Change Report"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14)
    ws_sum["A2"] = f"Generated: {run_date}  |  Comparing: {prev_folder} → {curr_folder}"
    ws_sum["A2"].font = Font(name="Arial", size=9, color="666666")

    sum_headers = [("Site", 24), ("Status", 16), ("# Changes", 12),
                   ("Pixel Diff", 12), ("Confidence", 13), ("Change Summary", 60)]
    for col_idx, (h, w) in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=4, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
    ws_sum.row_dimensions[4].height = 24
    ws_sum.freeze_panes = "A5"

    fill_changed = PatternFill("solid", start_color="FFF2CC")
    fill_ok      = PatternFill("solid", start_color="F0FFF0")
    fill_skip    = PatternFill("solid", start_color="F5F5F5")

    for row_offset, res in enumerate(results, 5):
        result = res["result"]
        has_changes = result.get("has_changes", False)
        skipped = result.get("skipped_claude", False)
        n_changes = len(result.get("changes", []))
        pixel_count = result.get("pixel_count", "N/A")
        row_fill = fill_changed if has_changes else (fill_skip if skipped else fill_ok)

        status = "🔴 CHANGED" if has_changes else ("⏭ No pixels changed" if skipped else "✅ No Change")
        vals = [res["site_name"], status, n_changes if has_changes else "",
                pixel_count, result.get("confidence", ""), result.get("change_summary", "")]
        for col_idx, val in enumerate(vals, 1):
            cell = ws_sum.cell(row=row_offset, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9); cell.border = border
            cell.fill = row_fill
            cell.alignment = left_wrap if col_idx == 6 else center
        ws_sum.row_dimensions[row_offset].height = 28

    # Changes Detail sheet
    ws_det = wb.create_sheet("Changes Detail")
    det_headers = [("Site", 24), ("Card Name", 32), ("Change Type", 20),
                   ("Previous Value", 30), ("Current Value", 30),
                   ("Severity", 10), ("Notes", 40), ("Confidence", 13)]
    for col_idx, (h, w) in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws_det.column_dimensions[get_column_letter(col_idx)].width = w
    ws_det.row_dimensions[1].height = 24
    ws_det.freeze_panes = "A2"

    det_row = 2
    for res in results:
        if not res["result"].get("has_changes"):
            continue
        for change in res["result"].get("changes", []):
            severity = change.get("severity", "low")
            sev_fill = PatternFill("solid", start_color={"high": "FFD7D7", "medium": "FFE9CC", "low": "D7FFD7"}.get(severity, "FFFFFF"))
            vals = [res["site_name"], change.get("card_name", ""),
                    CHANGE_TYPE_LABELS.get(change.get("change_type", "other"), change.get("change_type", "")),
                    change.get("previous_value", ""), change.get("current_value", ""),
                    severity.upper(), change.get("notes", ""), res["result"].get("confidence", "")]
            for col_idx, val in enumerate(vals, 1):
                cell = ws_det.cell(row=det_row, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=9); cell.border = border
                cell.fill = sev_fill if col_idx == 6 else None
                cell.alignment = left_wrap if col_idx in (3, 4, 5, 7) else center
            ws_det.row_dimensions[det_row].height = 32
            det_row += 1

    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(det_headers))}1"
    wb.save(CHANGE_REPORT_XLSX)
    print(f"✅ Change report → {CHANGE_REPORT_XLSX}")


def write_markdown_summary(results, run_date):
    lines = ["# Competitive Intelligence: Visual Change Report",
             f"**Generated:** {run_date}", "", "---", ""]
    changed = [r for r in results if r["result"].get("has_changes")]
    skipped = [r for r in results if r["result"].get("skipped_claude")]

    if not changed:
        lines.append("## ✅ No significant visual changes detected this month.")
    else:
        lines.append(f"## 🔴 {len(changed)} site(s) with changes\n")
        for res in changed:
            result = res["result"]
            lines.append(f"### {res['site_name']}")
            lines.append(f"**Summary:** {result.get('change_summary', '')}")
            lines.append(f"**Confidence:** {result.get('confidence', '')}  ")
            for change in result.get("changes", []):
                sev_icon = {"high": "🔴", "medium": "🟠", "low": "🟡"}.get(change.get("severity"), "⚪")
                lines.append(f"- {sev_icon} **{change.get('card_name', '')}** — {CHANGE_TYPE_LABELS.get(change.get('change_type','other'), '')}")
                lines.append(f"  - Before: `{change.get('previous_value', 'n/a')}`")
                lines.append(f"  - After: `{change.get('current_value', 'n/a')}`")
            lines.append("")

    lines.append("---")
    lines.append(f"## ⏭️ {len(skipped)} site(s) skipped (no pixel changes)")
    for res in skipped:
        lines.append(f"- ✅ {res['site_name']} — {res['result'].get('confidence_notes', '')}")

    with open(CHANGE_SUMMARY_MD, "w") as f:
        f.write("\n".join(lines))
    print(f"📄 Markdown summary → {CHANGE_SUMMARY_MD}")


def main():
    if not ANTHROPIC_API_KEY:
        print("❌ ANTHROPIC_API_KEY not set")
        sys.exit(1)

    prev_folder, curr_folder = find_screenshot_folders()
    if not curr_folder:
        print("❌ No screenshot folders found"); sys.exit(1)
    if not prev_folder:
        print("⚠️  Only one folder — nothing to compare. Run again next month.")
        sys.exit(0)

    print(f"📁 Previous: {prev_folder}")
    print(f"📁 Current:  {curr_folder}")

    pairs = match_screenshots(prev_folder, curr_folder)
    run_date = datetime.now().strftime("%Y-%m-%d")
    results = []
    claude_calls = 0
    skipped_calls = 0

    for site_name, prev_path, curr_path in pairs:
        print(f"\n→ {site_name}")
        if prev_path is None:
            results.append({"site_name": site_name, "status": "new_site",
                "result": {"has_changes": True, "change_summary": "New site added",
                    "changes": [{"card_name": "All", "change_type": "card_added",
                        "previous_value": "Not tracked", "current_value": "Now tracked",
                        "severity": "medium", "notes": ""}],
                    "confidence": "high", "confidence_notes": ""}})
            continue
        if curr_path is None:
            results.append({"site_name": site_name, "status": "removed_site",
                "result": {"has_changes": True, "change_summary": "Site missing",
                    "changes": [{"card_name": "All", "change_type": "offer_removed",
                        "previous_value": "Was tracked", "current_value": "Missing",
                        "severity": "high", "notes": "Check if URL changed"}],
                    "confidence": "high", "confidence_notes": ""}})
            continue

        result = compare_images(prev_path, curr_path)

        if result.get("skipped_claude"):
            skipped_calls += 1
        else:
            claude_calls += 1

        n_changes = len(result.get("changes", []))
        status = "🔴 CHANGED" if result.get("has_changes") else ("⏭ Skipped" if result.get("skipped_claude") else "✅ No change")
        print(f"  {status} — {n_changes} change(s)")
        results.append({"site_name": site_name, "status": "compared", "result": result})

    write_change_report(results, run_date, prev_folder, curr_folder)
    write_markdown_summary(results, run_date)

    changed_count = sum(1 for r in results if r["result"].get("has_changes"))
    print(f"\n{'='*60}")
    print(f"DONE: {changed_count}/{len(results)} sites with changes")
    print(f"Claude calls: {claude_calls} | Skipped (no pixel change): {skipped_calls}")
    estimated_cost = claude_calls * 0.07
    print(f"Estimated Claude cost: ~${estimated_cost:.2f}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
