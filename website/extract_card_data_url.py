"""
extract_card_data_url.py
========================
Option A: Playwright renders each card offer URL (JS-aware),
extracts visible text, sends to Claude Haiku for structured extraction.

Much more accurate than raw HTML requests — captures JS-rendered offers.

Outputs: data/card_data_url_YYYY_MM.xlsx + data/card_data_url_YYYY_MM.json
"""

import asyncio
import json
import os
import re
import sys
import time
from datetime import datetime

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import CARD_OFFER_SITES

CLAUDE_MODEL = "claude-haiku-4-5-20251001"
OUTPUT_DIR   = os.path.join(os.path.dirname(__file__), "data")
MAX_CHARS    = 10000  # more than raw HTML since text is cleaner

EXTRACTION_PROMPT = """You are extracting credit card offer data from a travel airline credit card webpage.

Site: {name}
URL: {url}

Page text (JS-rendered, visible content only):
{text}

Extract ALL credit cards visible. For each return a JSON object:
- card_name: full card name
- annual_fee: fee in dollars as number, 0 if none, null if not shown
- welcome_bonus_miles: bonus miles/points as number (e.g. 60000), null if not shown
- welcome_bonus_description: full bonus text (e.g. "60,000 miles after $3,000 in 3 months")
- spend_threshold: minimum spend in dollars as number, null if not shown
- spend_months: months to meet spend as number, null if not shown
- has_limited_time_offer: true if "limited time", "offer ends", "act now" language present
- limited_time_details: description if limited time offer, null otherwise
- earn_rate: primary earning rate (e.g. "3x miles on Delta"), null if not shown
- card_tier: "basic", "mid", "premium", or "business"

Return ONLY a JSON array. No markdown, no explanation.
If no cards found: []"""


# ─────────────────────────────────────────────
# CLAUDE SETUP
# ─────────────────────────────────────────────

_client = None

def get_client():
    global _client
    if _client:
        return _client
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("❌ ANTHROPIC_API_KEY not set")
        return None
    try:
        _client = anthropic.Anthropic(api_key=api_key)
        print(f"✅ Claude ready ({CLAUDE_MODEL})")
        return _client
    except Exception as e:
        print(f"❌ Claude init failed: {e}")
        return None


# ─────────────────────────────────────────────
# PLAYWRIGHT TEXT EXTRACTION
# ─────────────────────────────────────────────

async def fetch_rendered_text(page, site):
    """Use Playwright to render the page fully, then extract visible text."""
    try:
        print(f"  🌐 Loading page (wait: {site.get('wait_time', 8)}s)...")
        await page.goto(site["url"], wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(site.get("wait_time", 8))

        # Dismiss cookie banners
        for selector in [
            '[class*="cookie"]', '[id*="cookie"]', '[class*="consent"]',
            '#onetrust-banner-sdk', '[role="dialog"]',
        ]:
            try:
                for el in await page.locator(selector).all():
                    if await el.is_visible():
                        await el.evaluate('el => el.style.display = "none"')
            except:
                pass

        await asyncio.sleep(1)

        # Get visible text only — much cleaner than raw HTML
        text = await page.inner_text("body")
        text = re.sub(r"\s+", " ", text).strip()
        print(f"  📄 Rendered text: {len(text)} chars")
        return text[:MAX_CHARS]

    except Exception as e:
        print(f"  ⚠️  Page render failed: {e}")
        return ""


async def fetch_all_texts(sites):
    """Render all sites and return dict of site_name → text."""
    from playwright.async_api import async_playwright
    results = {}

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"]
        )
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = await context.new_page()

        for site in sites:
            print(f"\n  → {site['name']}")
            text = await fetch_rendered_text(page, site)
            results[site["name"]] = text

        await browser.close()

    return results


# ─────────────────────────────────────────────
# EXTRACTION
# ─────────────────────────────────────────────

def extract_from_text(site, text, client):
    """Send rendered page text to Claude for structured extraction."""
    if not text:
        return []

    prompt = EXTRACTION_PROMPT.format(
        name=site["name"], url=site["url"], text=text
    )
    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        cards = json.loads(raw)
        return cards if isinstance(cards, list) else []
    except Exception as e:
        print(f"  ⚠️  Extraction failed: {e}")
        return []


# ─────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────

def write_xlsx(records, run_date, method_label):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    month_str = run_date[:7].replace("-", "_")
    filename  = os.path.join(OUTPUT_DIR, f"card_data_url_{month_str}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = run_date[:7]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F3864")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    flag_fill   = PatternFill("solid", start_color="FFF2CC")
    alt_fill    = PatternFill("solid", start_color="F5F8FF")

    headers = [
        ("Site", 22), ("Company", 16), ("Card Name", 32), ("Tier", 10),
        ("Annual Fee", 13), ("Welcome Bonus Miles", 18), ("Spend Threshold ($)", 18),
        ("Spend Months", 13), ("Bonus Description", 42), ("Earn Rate", 30),
        ("Limited Time?", 14), ("Limited Time Details", 36),
        ("Method", 12), ("Source URL", 40), ("Run Date", 13),
    ]

    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, rec in enumerate(records, 2):
        is_alt   = row_idx % 2 == 0
        row_fill = flag_fill if rec.get("has_limited_time_offer") else (alt_fill if is_alt else None)
        values = [
            rec.get("site_name", ""), rec.get("company", ""), rec.get("card_name", ""),
            rec.get("card_tier", ""), rec.get("annual_fee"), rec.get("welcome_bonus_miles"),
            rec.get("spend_threshold"), rec.get("spend_months"),
            rec.get("welcome_bonus_description", ""), rec.get("earn_rate", ""),
            "✅ YES" if rec.get("has_limited_time_offer") else "No",
            rec.get("limited_time_details", ""), method_label,
            rec.get("source_url", ""), run_date,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border; cell.font = Font(name="Arial", size=9)
            cell.alignment = center if col in (4,5,6,7,8,11,13,15) else left
            if row_fill: cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 36

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(filename)
    print(f"\n  💾 Saved → {filename}")
    return filename


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    client   = get_client()
    run_date = datetime.now().strftime("%Y-%m-%d")
    records  = []

    print(f"\n{'='*60}")
    print(f"  Card Extractor — Option A: Playwright Rendered Text")
    print(f"  Sites: {len(CARD_OFFER_SITES)}  |  Date: {run_date}")
    print(f"{'='*60}\n")

    # Step 1 — render all pages with Playwright
    print("  📸 Rendering all pages...\n")
    text_map = asyncio.run(fetch_all_texts(CARD_OFFER_SITES))

    # Step 2 — extract card data from rendered text
    print(f"\n  🤖 Extracting card data with Claude...\n")
    for idx, site in enumerate(CARD_OFFER_SITES, 1):
        print(f"[{idx}/{len(CARD_OFFER_SITES)}] {site['name']}")

        if not client:
            records.append({"site_name": site["name"], "company": site["company"],
                            "card_name": "⚠️ No Claude", "source_url": site["url"]})
            continue

        text = text_map.get(site["name"], "")
        cards = extract_from_text(site, text, client)

        if not cards:
            print(f"  ⚠️  No cards extracted")
            records.append({"site_name": site["name"], "company": site["company"],
                            "card_name": "⚠️ No data", "source_url": site["url"]})
        else:
            print(f"  ✅ {len(cards)} card(s)")
            for card in cards:
                card["site_name"]  = site["name"]
                card["company"]    = site["company"]
                card["source_url"] = site["url"]
                records.append(card)

        time.sleep(0.5)

    write_xlsx(records, run_date, method_label="Playwright+Text")

    json_path = os.path.join(OUTPUT_DIR, f"card_data_url_{run_date}.json")
    with open(json_path, "w") as f:
        json.dump(records, f, indent=2)
    print(f"  📄 JSON → {json_path}")
    print(f"\n{'='*60}")
    print(f"  ✅ Done! {len(records)} total records")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
